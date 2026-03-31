"""
export_status.py — CESTRIN Peek Monitor
========================================
Citeste 0_Centralizator_PEEK-VEK.xlsx si trimite datele la platforma web Flask.

CUM SE RULEAZA:
  1. Porniti mai intai serverul:  python app.py
  2. Apoi rulati acest script:    python export_status.py
"""

import json, sys, requests
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("Lipsa openpyxl. Rulati: pip install openpyxl requests")
    sys.exit(1)

# ══════════════════════════════════════════════════════════════════════════════
#  CONFIGURARE — editati doar aceste 3 valori
# ══════════════════════════════════════════════════════════════════════════════

EXCEL_PATH = r"L:\BIDMRCT\datePEEK\0_Centralizator_PEEK-VEK.xlsx"
SERVER_URL = "http://localhost:5000/api/upload-status"
API_KEY    = "SCHIMBA_KEY_SECRET"

# ══════════════════════════════════════════════════════════════════════════════

MAPARE_STARI = {
    "clasificator":     "clasificator",
    "totalizator":      "totalizator",
    "defect":           "defect",
    "nu functioneaza":  "nefunctional",
    "nu functioneaza":  "nefunctional",
    "fara conexiune":   "fara_conexiune",
    "fara conexiune":   "fara_conexiune",
}

EMOJI = {
    "clasificator":"green","totalizator":"yellow",
    "defect":"orange","nefunctional":"red",
    "fara_conexiune":"purple","unknown":"white"
}

def norm(val):
    if not val: return "unknown"
    return MAPARE_STARI.get(str(val).strip().lower(), "unknown")

def idx(header, col):
    try: return header.index(col)
    except ValueError: return None

def citeste_excel(cale):
    path = Path(cale)
    if not path.exists():
        print(f"\nFisierul Excel nu a fost gasit:\n   {path}")
        print("Verificati valoarea EXCEL_PATH din acest script.")
        sys.exit(1)

    print(f"Citesc: {path.name}")
    wb = load_workbook(path, read_only=True, data_only=True)

    # Sheet Contoare
    ws_c    = wb["Contoare"]
    rows_c  = list(ws_c.iter_rows(values_only=True))
    hdr_c   = [str(v).strip() if v else "" for v in rows_c[0]]

    i_cid  = idx(hdr_c, "Contor")
    i_drum = idx(hdr_c, "Drum")
    i_km   = idx(hdr_c, "Pozitie km") or idx(hdr_c, "Pozi\u021bie km")
    i_loc  = idx(hdr_c, "Localitate")
    i_tip  = idx(hdr_c, "Tip")
    i_ip   = idx(hdr_c, "IP")
    i_x    = idx(hdr_c, "x")
    i_y    = idx(hdr_c, "y")

    # fallback pentru "Poziție km" cu diacritice
    if i_km is None:
        for i, h in enumerate(hdr_c):
            if "km" in h.lower() or "pozi" in h.lower():
                i_km = i; break

    contoare = {}
    erori    = []
    for row in rows_c[1:]:
        if not row or row[i_cid] is None: continue
        cid = str(row[i_cid]).strip()
        if not cid or cid.startswith('#') or (not cid[0].isdigit() and not cid[0].isalpha()):
            erori.append(cid); continue
        # Ignora randul cu eroare CampiaturziiCJ / #N/A
        drum_val = row[i_drum] if i_drum is not None else None
        if drum_val and str(drum_val).startswith('#'):
            erori.append(cid); continue

        def v(i):
            return row[i] if i is not None and i < len(row) else None

        contoare[cid] = {
            "id":         cid,
            "drum":       str(v(i_drum)  or "").strip(),
            "km":         str(v(i_km)    or "").strip(),
            "localitate": str(v(i_loc)   or "").strip(),
            "tip":        str(v(i_tip)   or "").strip(),
            "ip":         str(v(i_ip)    or "").strip(),
            "x":          v(i_x),
            "y":          v(i_y),
        }

    print(f"  [Contoare] {len(contoare)} posturi valide" +
          (f", {len(erori)} ignorate" if erori else ""))

    # Sheet Media Zilnica Lunara
    ws_m    = wb["Media Zilnica Lunara"]
    rows_m  = list(ws_m.iter_rows(values_only=True))
    hdr_m   = [str(v).strip() if v else "" for v in rows_m[1]]  # randul 2 = header

    j_cid  = idx(hdr_m, "Contor")
    j_an   = idx(hdr_m, "An")
    j_luna = idx(hdr_m, "Luna")
    j_mod  = idx(hdr_m, "Mod de functionare")
    j_tot  = idx(hdr_m, "Total")
    j_zile = idx(hdr_m, "Zile cu \xeenregistr\u0103ri") or idx(hdr_m, "Zile cu inregistrari")

    # fallback zile
    if j_zile is None:
        for i, h in enumerate(hdr_m):
            if "zile" in h.lower(): j_zile = i; break

    if j_mod is None:
        print(f"  ATENTIE: Coloana 'Mod de functionare' nu a fost gasita!")
        print(f"  Coloane detectate: {hdr_m}")

    recent = {}
    for row in rows_m[2:]:
        if not row or row[j_cid] is None: continue
        cid  = str(row[j_cid]).strip()
        an   = row[j_an]
        luna = row[j_luna]
        if an is None or luna is None: continue
        mod  = row[j_mod]  if j_mod  is not None else None
        tot  = row[j_tot]  if j_tot  is not None else None
        zile = row[j_zile] if j_zile is not None else None
        if cid not in recent or (int(an), int(luna)) > recent[cid][:2]:
            recent[cid] = (int(an), int(luna), mod, tot, zile)

    print(f"  [Media] {len(recent)} contoare cu date recente")

    # Combinare
    posturi = []
    fara_xy = 0
    for cid, info in contoare.items():
        rec = recent.get(cid)
        if rec:
            an, luna, mod, tot, zile = rec
            stare       = norm(mod)
            mod_str     = str(mod).strip() if mod else "—"
            ultima_luna = f"{an}-{luna:02d}"
        else:
            stare = "unknown"; mod_str = "—"
            ultima_luna = None; tot = None; zile = None

        if not info["x"] or not info["y"]:
            fara_xy += 1

        posturi.append({
            "id":              cid,
            "drum":            info["drum"],
            "km":              info["km"],
            "localitate":      info["localitate"],
            "tip":             info["tip"],
            "stare":           stare,
            "mod_functionare": mod_str,
            "ultima_luna":     ultima_luna,
            "total_vehicule":  tot,
            "zile_inreg":      str(zile) if zile else None,
            "x":               info["x"],
            "y":               info["y"],
        })
        icons = {"clasificator":"[V]","totalizator":"[G]","defect":"[P]",
                 "nefunctional":"[R]","fara_conexiune":"[M]","unknown":"[ ]"}
        print(f"  {icons.get(stare,'[?]')} {cid:6s}  {info['drum']:8s}  "
              f"km {info['km']:10s}  {info['localitate']:22s}  {mod_str}")

    print(f"\n  Total: {len(posturi)} posturi")
    if fara_xy:
        print(f"  {fara_xy} fara X/Y — plasati manual pe harta dupa upload")
    return posturi


def trimite(posturi):
    print(f"\nTriimit {len(posturi)} posturi la {SERVER_URL} ...")
    try:
        r = requests.post(
            SERVER_URL,
            json={"contoare": posturi, "sursa": "export_cestrin"},
            headers={
                "X-API-Key": API_KEY,
                "User-Agent": "CESTRIN-Exporter/1.0"
            },
            timeout=30,
        )
        if r.status_code == 200:
            print(f"SUCCES: {r.json().get('posturi','?')} posturi inregistrate.")
            print("Deschideti harta in browser — contoarele sunt acum vizibile.")
            return True
        elif r.status_code == 403:
            print("Cheie API invalida (403).")
            print("Verificati: API_KEY din script = UPLOAD_API_KEY din app.py")
        elif r.status_code == 404:
            print("Endpoint negasit (404). Flask ruleaza?")
        else:
            print(f"Eroare {r.status_code}: {r.text[:300]}")
    except requests.exceptions.ConnectionError:
        print(f"Nu se poate conecta la {SERVER_URL}")
        print("Verificati ca Flask ruleaza:  python app.py")
    except Exception as e:
        print(f"Eroare: {e}")
    return False


if __name__ == "__main__":
    print("=" * 65)
    print(f"  CESTRIN Export Status — {datetime.now():%Y-%m-%d %H:%M:%S}")
    print("=" * 65)

    posturi = citeste_excel(EXCEL_PATH)
    if not posturi: sys.exit(1)

    Path("status_last_export.json").write_text(
        json.dumps({"contoare": posturi}, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    print("\nBackup local: status_last_export.json")

    ok = trimite(posturi)
    print("\n" + ("SUCCES" if ok else "ESUAT"))
    sys.exit(0 if ok else 1)
