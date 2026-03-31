"""
export_status.py — Rulează pe serverul/PC-ul CESTRIN
─────────────────────────────────────────────────────
Citește centralizatorul Excel (toate DRDP-urile),
geocodează automat după DN + km + localitate,
trimite datele la platforma web.

Rulare manuală : python export_status.py
Task Scheduler : săptămânal după procesarea .bin
"""

import json, sys, time
from pathlib import Path
from datetime import datetime

try:
    import requests
    from openpyxl import load_workbook
except ImportError:
    print("pip install openpyxl requests"); sys.exit(1)

# ══════════════════════════════════════════════════════
# CONFIGURARE
# ══════════════════════════════════════════════════════
CFG = {
    "server_url":  "https://trafic-cestrin.ro/api/upload-status",
    "api_key":     "SCHIMBA_CU_UPLOAD_API_KEY_DIN_ENV",

    # Calea centralizatorului Excel cu toate DRDP-urile
    "excel_path":  r"C:\CESTRIN\Date_Trafic\centralizator_national.xlsx",

    # Coloane Excel (litera coloanei)
    "col_id":      "A",   # ID post: P001, P002...
    "col_drdp":    "B",   # DRDP: Cluj, Timișoara, Iași...
    "col_drum":    "C",   # DN1, DN7...
    "col_km":      "D",   # număr km
    "col_localit": "E",   # Localitate
    "col_judet":   "F",   # Județ (pentru geocoding mai precis)
    "col_benzi":   "G",   # nr benzi
    "col_tip":     "H",   # Peek TD8 / Peek Trax
    "col_stare":   "I",   # ← starea de funcționare
    "col_desc":    "J",   # data ultimei descărcări
    "col_nume":    "K",   # nume complet post (opțional)

    "start_row":   2,     # primul rând cu date (după header)

    # Mapare valori Excel → stări platformă
    "mapare_stari": {
        "functional":     "ok",    "funcțional":    "ok",
        "ok":             "ok",    "activ":         "ok",    "1":  "ok",
        "partial":        "partial","parțial":       "partial",
        "defect partial": "partial","banda lipsa":   "partial",
        "offline":        "offline","defect":        "offline",
        "nefunctional":   "offline","nefuncțional":  "offline",
        "inactiv":        "offline","0":             "offline",
        "mentenanta":     "maintenance","mentenanță":"maintenance",
        "reparatie":      "maintenance","reparație":  "maintenance",
    },
}


def col_idx(c): 
    r=0
    for ch in c.upper().strip(): r=r*26+(ord(ch)-64)
    return r-1

def norm_stare(v):
    if v is None: return "unknown"
    return CFG["mapare_stari"].get(str(v).strip().lower(), "unknown")


def citeste_excel():
    p = Path(CFG["excel_path"])
    if not p.exists():
        print(f"❌ Fișier negăsit: {p}"); sys.exit(1)

    print(f"📂 {p.name}")
    wb = load_workbook(p, data_only=True)

    # Caută foaia principală
    ws = None
    for name in ['Centralizator','centralizator','National','Sheet1','Foaie1']:
        if name in wb.sheetnames: ws=wb[name]; break
    if not ws: ws = wb.active
    print(f"   Foaie: {ws.title}")

    idx = {k: col_idx(CFG[f'col_{k}'])
           for k in ['id','drdp','drum','km','localit','judet','benzi','tip','stare','desc','nume']}

    posturi, erori = [], 0
    for row in ws.iter_rows(min_row=CFG['start_row'], values_only=True):
        pid = str(row[idx['id']] or '').strip()
        if not pid: continue

        stare = norm_stare(row[idx['stare']])

        desc_v = row[idx['desc']]
        if isinstance(desc_v, datetime): ultima = desc_v.strftime('%Y-%m-%d %H:%M')
        elif desc_v: ultima = str(desc_v).strip()
        else: ultima = None

        post = {
            'id':         pid,
            'drdp':       str(row[idx['drdp']]  or '').strip(),
            'drum':       str(row[idx['drum']]  or '').strip(),
            'km':         row[idx['km']],
            'localitate': str(row[idx['localit']] or '').strip(),
            'judet':      str(row[idx['judet']]  or '').strip(),
            'benzi':      row[idx['benzi']],
            'tip':        str(row[idx['tip']]   or '').strip(),
            'stare':      stare,
            'ultima_desc':ultima,
            'nume':       str(row[idx['nume']]  or '').strip(),
            # lat/lng vor fi completate de geocoder pe server
        }
        posturi.append(post)
        ic = {'ok':'✅','partial':'⚠️','offline':'❌','maintenance':'🔧','unknown':'❓'}[stare]
        print(f"   {ic} {pid:8s} {post['drdp']:12s} {post['drum']:8s} km {str(post['km'] or '?'):5s}")

    print(f"\n   Total: {len(posturi)} posturi ({erori} erori)")
    return posturi


def trimite(posturi):
    payload = {
        'contoare': posturi,
        'sursa':    'export_status_cestrin',
        'versiune': '2.0',
    }
    print(f"\n🌐 Trimit la {CFG['server_url']} ...")
    try:
        r = requests.post(
            CFG['server_url'], json=payload,
            headers={'X-API-Key': CFG['api_key']},
            timeout=60,
        )
        if r.status_code == 200:
            d = r.json()
            print(f"✅ Succes! {d.get('posturi','?')} posturi înregistrate.")
            return True
        else:
            print(f"❌ Server: {r.status_code} — {r.text[:200]}")
    except requests.exceptions.ConnectionError:
        print("❌ Nu se poate conecta la server.")
    except requests.exceptions.Timeout:
        print("❌ Timeout (>60s).")
    except Exception as e:
        print(f"❌ {e}")
    return False


if __name__ == '__main__':
    print("=" * 55)
    print(f"  CESTRIN Export Status — {datetime.now():%Y-%m-%d %H:%M:%S}")
    print("=" * 55 + "\n")

    posturi = citeste_excel()
    if not posturi: sys.exit(1)

    # Backup local
    with open('status_last_export.json','w',encoding='utf-8') as f:
        json.dump({'contoare':posturi},f,ensure_ascii=False,indent=2)
    print("💾 Backup local: status_last_export.json")

    ok = trimite(posturi)
    print("\n" + "="*55)
    print(f"  {'✅ SUCCES' if ok else '❌ EȘUAT'}")
    print("="*55)
    sys.exit(0 if ok else 1)
